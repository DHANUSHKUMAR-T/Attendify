"""
Student Attendance Management System
=====================================
Flask application with MongoDB backend
Handles registration, login, and attendance tracking
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
import os
import io
import random
import openpyxl
from openpyxl.styles import Font
from datetime import date, datetime
from functools import wraps
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from pymongo import MongoClient
from bson import ObjectId

# ─────────────────────────────────────────
# App Configuration
# ─────────────────────────────────────────
app = Flask(__name__)
app.secret_key = "sams_secret_key_2024"   # Change this in production!

# MongoDB connection – use environment variable MONGO_URI
MONGO_URI="mongodb+srv://tdhanu:Dhanu123@clustersumma.ggzzczb.mongodb.net/attendance_db?retryWrites=true&w=majority&appName=ClusterSumma"
client = MongoClient(MONGO_URI)
db = client["attendance_db"]
students_collection = db["students"]
attendance_collection = db["attendance"]

# Ensure unique index on register_number
students_collection.create_index("register_number", unique=True)

# Strict Attendance Settings
ADMIN_PASSWORD = "admin123"
ADMIN_CLEAR_SECRET = "CLEAR2024"

# Global PIN for attendance marking
CURRENT_PIN = None

# Google Sheets Configuration
SHEET_URL = "https://docs.google.com/spreadsheets/d/1Thkg9MvtngFwsOD6PadpgJe7C7RbVYz2dfUW_3CgyCc/edit"
CREDENTIALS_FILE = os.path.join(os.path.dirname(__file__), "service_account.json")

def sync_to_google_sheets(data):
    """Sync attendance data to Google Sheets.
    data: dict with date, time, reg_no, name, course, status
    """
    if not os.path.exists(CREDENTIALS_FILE):
        print("Google Sheets Sync: service_account.json not found. Skipping sync.")
        return False
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, scope)
        client_gs = gspread.authorize(creds)
        sheet = client_gs.open_by_url(SHEET_URL).sheet1
        sheet.append_row([
            data.get('date'),
            data.get('time'),
            data.get('reg_no'),
            data.get('name'),
            data.get('course'),
            data.get('status')
        ])
        return True
    except Exception as e:
        print(f"Google Sheets Sync Error: {e}")
        return False

# ─────────────────────────────────────────
# Login Required Decorator
# ─────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "student_id" not in session:
            flash("Please log in to continue.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            flash("Please log in as Admin to continue.", "warning")
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

# ─────────────────────────────────────────
# Routes
# ─────────────────────────────────────────

@app.route("/")
def index():
    return redirect(url_for("login"))

# ── Registration ──────────────────────────
@app.route("/register", methods=["GET", "POST"]) 
def register():
    if request.method == "POST":
        reg_no = request.form.get("register_number", "").strip().upper()
        name = request.form.get("name", "").strip()
        course = request.form.get("course", "").strip().upper()
        mobile = request.form.get("mobile", "").strip()
        alt_mob = request.form.get("alt_mobile", "").strip()
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")

        errors = []
        if not reg_no:
            errors.append("Register number is required.")
        if not name:
            errors.append("Name is required.")
        if not course:
            errors.append("Course is required.")
        if not mobile.isdigit() or len(mobile) != 10:
            errors.append("Enter a valid 10-digit mobile number.")
        if alt_mob and (not alt_mob.isdigit() or len(alt_mob) != 10):
            errors.append("Alternative mobile must be a valid 10-digit number.")
        if len(password) < 6:
            errors.append("Password must be at least 6 characters.")
        if password != confirm:
            errors.append("Passwords do not match.")
        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("register.html", form=request.form)

        hashed_pw = generate_password_hash(password)
        try:
            students_collection.insert_one({
                "register_number": reg_no,
                "name": name,
                "course": course,
                "mobile": mobile,
                "alt_mobile": alt_mob or None,
                "password": hashed_pw
            })
            flash("Registration successful! Please log in.", "success")
            return redirect(url_for("login"))
        except Exception as e:
            flash("Register number already exists or database error.", "danger")
            return render_template("register.html", form=request.form)
    return render_template("register.html", form={})

# ── Login ─────────────────────────────────
@app.route("/login", methods=["GET", "POST"]) 
def login():
    if "student_id" in session:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        reg_no = request.form.get("register_number", "").strip().upper()
        password = request.form.get("password", "")
        student = students_collection.find_one({"register_number": reg_no})
        if student and check_password_hash(student["password"], password):
            session["student_id"] = str(student["_id"])
            session["student_name"] = student["name"]
            session["register_number"] = student["register_number"]
            session["course"] = student["course"]
            flash(f"Welcome back, {student['name']}!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid register number or password.", "danger")
    return render_template("login.html")

# ── Logout ────────────────────────────────
@app.route("/logout")
@login_required
def logout():
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for("login"))

# ── Dashboard ─────────────────────────────
@app.route("/dashboard")
@login_required
def dashboard():
    student_id = session["student_id"]
    today = date.today().isoformat()

    # Attendance for today
    already_marked_doc = attendance_collection.find_one({"student_id": ObjectId(student_id), "date": today})
    already_marked = already_marked_doc["status"] if already_marked_doc else None

    # Full history (newest first)
    history_cursor = attendance_collection.find({"student_id": ObjectId(student_id)}).sort("date", -1)
    history = list(history_cursor)

    total = len(history)
    present = sum(1 for r in history if r.get("status") == "Present")
    absent = total - present
    pct = round((present / total * 100), 1) if total > 0 else 0

    student = students_collection.find_one({"_id": ObjectId(student_id)})

    return render_template(
        "dashboard.html",
        today=today,
        already_marked=already_marked,
        history=history,
        total=total,
        present=present,
        absent=absent,
        percentage=pct,
        student=student
    )

# ── Mark Attendance ───────────────────────
@app.route("/mark_attendance", methods=["POST"])
@login_required
def mark_attendance():
    global CURRENT_PIN
    student_id = session["student_id"]
    today = date.today().isoformat()
    status = request.form.get("status")
    user_pin = request.form.get("otp_pin")

    if CURRENT_PIN and user_pin != str(CURRENT_PIN):
        flash("Invalid OTP PIN. Please check with your teacher.", "danger")
        return redirect(url_for("dashboard"))
    if status not in ("Present", "Absent"):
        flash("Invalid status selection.", "danger")
        return redirect(url_for("dashboard"))

    try:
        attendance_collection.insert_one({
            "student_id": ObjectId(student_id),
            "date": today,
            "status": status
        })
        # Sync to Google Sheets
        student = students_collection.find_one({"_id": ObjectId(student_id)})
        if student:
            sync_data = {
                "date": today,
                "time": datetime.now().strftime("%H:%M:%S"),
                "reg_no": student["register_number"],
                "name": student["name"],
                "course": student["course"],
                "status": status
            }
            sync_to_google_sheets(sync_data)
        flash(f"Attendance marked as {status} for today ({today}).", "success")
    except Exception as e:
        flash("Attendance for today has already been recorded or an error occurred.", "warning")
    return redirect(url_for("dashboard"))

# ── Admin Routes ──────────────────────────
@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    if session.get("is_admin"):
        return redirect(url_for("admin_dashboard"))
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == ADMIN_PASSWORD:
            session["is_admin"] = True
            flash("Admin login successful.", "success")
            return redirect(url_for("admin_dashboard"))
        else:
            flash("Invalid admin password.", "danger")
    return render_template("admin_login.html")

@app.route("/admin_logout")
def admin_logout():
    session.pop("is_admin", None)
    flash("Admin logged out.", "info")
    return redirect(url_for("admin_login"))

@app.route("/admin_dashboard")
@admin_required
def admin_dashboard():
    target_date = request.args.get("date", date.today().isoformat())
    students = list(students_collection.find({}, {"register_number": 1, "name": 1, "course": 1, "balance": 1}))
    attendance_records = list(attendance_collection.find({"date": target_date}, {"student_id": 1, "status": 1}))
    att_map = {str(rec["student_id"]): rec["status"] for rec in attendance_records}
    student_data = []
    for s in students:
        sid = str(s["_id"])
        student_data.append({
            "id": sid,
            "register_number": s["register_number"],
            "name": s["name"],
            "course": s["course"],
            "balance": s.get("balance", 0),
            "status": att_map.get(sid, "Not Marked")
        })
    return render_template("admin_dashboard.html", students=student_data, target_date=target_date, current_pin=CURRENT_PIN)

@app.route("/admin/update_attendance", methods=["POST"])
@admin_required
def update_attendance():
    student_id = request.form.get("student_id")
    target_date = request.form.get("date")
    new_status = request.form.get("status")
    if new_status not in ("Present", "Absent"):
        flash("Invalid status.", "danger")
        return redirect(url_for("admin_dashboard", date=target_date))
    existing = attendance_collection.find_one({"student_id": ObjectId(student_id), "date": target_date})
    if existing:
        attendance_collection.update_one({"_id": existing["_id"]}, {"$set": {"status": new_status}})
    else:
        attendance_collection.insert_one({"student_id": ObjectId(student_id), "date": target_date, "status": new_status})
    # Sync to Google Sheets
    student = students_collection.find_one({"_id": ObjectId(student_id)})
    if student:
        sync_data = {
            "date": target_date,
            "time": datetime.now().strftime("%H:%M:%S"),
            "reg_no": student["register_number"],
            "name": student["name"],
            "course": student["course"],
            "status": new_status
        }
        sync_to_google_sheets(sync_data)
    flash(f"Attendance updated to {new_status}.", "success")
    return redirect(url_for("admin_dashboard", date=target_date))

@app.route("/admin/add_student", methods=["POST"])
@admin_required
def admin_add_student():
    reg_no = request.form.get("register_number", "").strip().upper()
    name = request.form.get("name", "").strip()
    course = request.form.get("course", "").strip().upper()
    mobile = request.form.get("mobile", "").strip()
    password = request.form.get("password", "123456")
    if not reg_no or not name:
        flash("Register number and name are required.", "danger")
        return redirect(url_for("admin_dashboard"))
    hashed_pw = generate_password_hash(password)
    try:
        students_collection.insert_one({
            "register_number": reg_no,
            "name": name,
            "course": course,
            "mobile": mobile,
            "password": hashed_pw
        })
        flash(f"Student {name} added successfully.", "success")
    except Exception as e:
        flash("Register number already exists.", "danger")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/delete_student/<string:student_id>", methods=["POST"])
@admin_required
def admin_delete_student(student_id):
    attendance_collection.delete_many({"student_id": ObjectId(student_id)})
    students_collection.delete_one({"_id": ObjectId(student_id)})
    flash("Student and their attendance records deleted.", "info")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/generate_pin", methods=["POST"])
@admin_required
def generate_pin():
    global CURRENT_PIN
    CURRENT_PIN = random.randint(1000, 9999)
    flash(f"New OTP PIN generated: {CURRENT_PIN}", "success")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/export_excel")
@admin_required
def export_excel():
    from openpyxl.styles import PatternFill
    
    # 1. Fetch all unique dates sorted chronologically
    sorted_dates = sorted(list(attendance_collection.distinct("date")))
    
    # 2. Fetch all students sorted by register number
    students = list(students_collection.find({}, {"register_number": 1, "name": 1, "course": 1, "balance": 1}).sort("register_number", 1))
    
    # 3. Fetch all attendance records
    attendance_records = list(attendance_collection.find())
    
    # 4. Map (student_id_str, date) -> status
    att_map = {}
    for r in attendance_records:
        att_map[(str(r["student_id"]), r["date"])] = r["status"]
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Attendance"
    
    # Headers
    headers = ["Register Number", "Name", "Course", "Outstanding Balance"]
    formatted_dates = []
    for d in sorted_dates:
        try:
            fd = datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            fd = d
        formatted_dates.append(fd)
        
    headers.extend(formatted_dates)
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    # Soft fills for statuses
    present_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    not_marked_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    
    present_font = Font(name="Inter", size=10, color="065F46")
    absent_font = Font(name="Inter", size=10, color="991B1B")
    default_font = Font(name="Inter", size=10, color="1F2937")
    
    # Populate student rows
    row_num = 2
    for s in students:
        sid = str(s["_id"])
        row_data = [
            s.get("register_number"),
            s.get("name"),
            s.get("course"),
            s.get("balance", 0)
        ]
        
        # Append status for each date
        for d in sorted_dates:
            status = att_map.get((sid, d), "Not Marked")
            row_data.append(status)
            
        ws.append(row_data)
        
        # Style cells in the row
        ws.cell(row=row_num, column=1).font = Font(name="Inter", size=10, bold=True)
        ws.cell(row=row_num, column=2).font = default_font
        ws.cell(row=row_num, column=3).font = default_font
        
        # Style Balance cell
        bal_cell = ws.cell(row=row_num, column=4)
        bal_val = row_data[3]
        bal_cell.font = Font(name="Inter", size=10, bold=True, color="991B1B" if bal_val > 0 else "065F46")
        
        # Style daily status cells
        for col_idx, d in enumerate(sorted_dates, start=5):
            cell = ws.cell(row=row_num, column=col_idx)
            status = cell.value
            if status == "Present":
                cell.fill = present_fill
                cell.font = present_font
            elif status == "Absent":
                cell.fill = absent_fill
                cell.font = absent_font
            else:
                cell.fill = not_marked_fill
                cell.font = default_font
                
        row_num += 1
        
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 22
    
    for col_idx in range(5, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 15
        
    ws.views.sheetView[0].showGridLines = True
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Master_Attendance_Report_{date.today().isoformat()}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin/clear_db", methods=["POST"])
@admin_required
def clear_db():
    pw1 = request.form.get("password1")
    pw2 = request.form.get("password2")
    confirm = request.form.get("confirm")
    if pw1 == ADMIN_PASSWORD and pw2 == ADMIN_CLEAR_SECRET and confirm == "yes":
        attendance_collection.delete_many({})
        students_collection.delete_many({})
        flash("Database cleared successfully (all students and attendance deleted).", "success")
    else:
        flash("Incorrect passwords or confirmation not provided. Action aborted.", "danger")
    return redirect(url_for("admin_dashboard"))

@app.route("/admin/manage_money/<string:student_id>", methods=["GET", "POST"])
@admin_required
def manage_money(student_id):
    student = students_collection.find_one({"_id": ObjectId(student_id)})
    if not student:
        flash("Student not found.", "danger")
        return redirect(url_for("admin_dashboard"))

    if request.method == "POST":
        action = request.form.get("action")
        prev_balance = float(student.get("balance", 0))

        if action == "direct_update":
            try:
                new_balance = float(request.form.get("new_balance", 0))
                note = request.form.get("note", "Manual balance adjustment").strip() or "Manual balance adjustment"
                change = new_balance - prev_balance
                
                transaction = {
                    "date": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "type": "Direct Adjustment",
                    "amount": change,
                    "prev_balance": prev_balance,
                    "new_balance": new_balance,
                    "note": note
                }
                
                students_collection.update_one(
                    {"_id": ObjectId(student_id)},
                    {
                        "$set": {"balance": new_balance},
                        "$push": {"payment_history": transaction}
                    }
                )
                flash(f"Balance directly updated to {new_balance}.", "success")
            except Exception as e:
                flash(f"Error updating balance: {e}", "danger")

        elif action == "record_payment":
            try:
                amount_given = float(request.form.get("amount_given", 0))
                payment_note = request.form.get("payment_note", "Fees payment").strip() or "Fees payment"
                new_balance = prev_balance - amount_given
                
                transaction = {
                    "date": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "type": "Payment Received",
                    "amount": amount_given,
                    "prev_balance": prev_balance,
                    "new_balance": new_balance,
                    "note": payment_note
                }
                
                students_collection.update_one(
                    {"_id": ObjectId(student_id)},
                    {
                        "$set": {"balance": new_balance},
                        "$push": {"payment_history": transaction}
                    }
                )
                flash(f"Recorded payment of {amount_given}. Outstanding balance is now {new_balance}.", "success")
            except Exception as e:
                flash(f"Error recording payment: {e}", "danger")

        # Refetch student
        student = students_collection.find_one({"_id": ObjectId(student_id)})

    return render_template("admin_manage_money.html", student=student)

# ─────────────────────────────────────────
# Entry Point
# ─────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True)
